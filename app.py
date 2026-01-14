import time
import requests
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook

# -----------------------
# Конфигурация
# -----------------------
SPEEDY_BASE_URL = "https://api.speedyindex.com/v2"

# -----------------------
# Вспомогательные функции
# -----------------------
def get_headers(api_key):
    return {
        "Authorization": api_key,
        "Content-Type": "application/json"
    }

def get_balance(api_key):
    """Получаем баланс аккаунта (Checker)"""
    try:
        url = f"{SPEEDY_BASE_URL}/account"
        resp = requests.get(url, headers=get_headers(api_key), timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            return data.get("balance", {}).get("checker", 0)
    except Exception:
        return None
    return None

def send_slack_notification(token, channel, message):
    """Отправка уведомления в Slack"""
    url = "https://slack.com/api/chat.postMessage"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    payload = {
        "channel": channel,
        "text": message
    }
    try:
        requests.post(url, headers=headers, json=payload, timeout=5)
    except Exception as e:
        print(f"Slack error: {e}")

def find_header_row(ws, max_scan=20):
    """Ищем строку заголовков (Referring Page URL)"""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        val = ws.cell(row=r, column=2).value
        if isinstance(val, str) and "referring page url" in val.lower():
            return r
    return 1

def looks_like_url(val):
    if not isinstance(val, str):
        return False
    s = val.strip().lower()
    return s.startswith("http://") or s.startswith("https://")

# Эту функцию кэшируем, чтобы не перечитывать тяжелый файл при кликах по интерфейсу
@st.cache_resource(ttl="1h", show_spinner=False)
def load_workbook_cached(file_content):
    return load_workbook(BytesIO(file_content))

# -----------------------
# Основной UI Streamlit
# -----------------------
st.set_page_config(page_title="SpeedyIndex Checker", layout="wide")
st.title("Проверка индексации (SpeedyIndex)")

# 1. Проверка Secrets
if "speedyindex" not in st.secrets or "slack" not in st.secrets:
    st.error("Ошибка конфигурации! Проверьте .streamlit/secrets.toml (секции [speedyindex] и [slack]).")
    st.stop()

api_key = st.secrets["speedyindex"]["api_key"]
slack_token = st.secrets["slack"]["bot_token"]
slack_channel = st.secrets["slack"]["channel_id"]

# 2. Отображение баланса
balance = get_balance(api_key)
col_bal, col_dummy = st.columns([1, 3])
with col_bal:
    if balance is not None:
        if balance > 1000:
            st.success(f"💰 Баланс Checker: **{balance}**")
        else:
            st.warning(f"💰 Баланс Checker: **{balance}** (мало!)")
    else:
        st.error("Не удалось получить баланс API")

st.markdown("---")

# 3. Загрузка файла
uploaded_file = st.file_uploader("Загрузите файл .xlsx", type=["xlsx"])

if uploaded_file:
    # --- БЛОК ЗАГРУЗКИ С ИНДИКАЦИЕЙ ---
    # Мы используем st.status, чтобы пользователь видел процесс
    with st.status("Чтение файла...", expanded=True) as status:
        st.write("Загрузка структуры Excel (это может занять время для больших файлов)...")
        try:
            # Загружаем через кэшируемую функцию
            # Важно: мы передаем bytes, чтобы кэш работал корректно
            wb_source = load_workbook_cached(uploaded_file.getvalue())
            
            # ВАЖНО: Кэшированный объект нельзя менять напрямую, если мы хотим
            # чистые данные при повторном запуске.
            # Но так как openpyxl copy долгий, мы будем аккуратны.
            # Для простоты: берем имена листов из кэша, а для обработки 
            # будем использовать этот же объект (но учтите, что он сохранится в памяти измененным до перезагрузки кэша)
            
            status.update(label="Файл успешно прочитан! ✅", state="complete", expanded=False)
        except Exception as e:
            st.error(f"Ошибка при чтении файла: {e}")
            st.stop()
    # -----------------------------------

    all_sheet_names = wb_source.sheetnames
    selected_sheets = []

    # Логика выбора листов
    if len(all_sheet_names) > 1:
        st.info(f"Найдено листов: {len(all_sheet_names)}")
        selected_sheets = st.multiselect(
            "Выберите листы для обработки:", 
            options=all_sheet_names,
            default=all_sheet_names
        )
    else:
        selected_sheets = all_sheet_names

    if not selected_sheets:
        st.warning("Выберите хотя бы один лист.")
        st.stop()

    # Кнопка запуска
    if st.button("🚀 Начать проверку"):
        
        # Чтобы не портить кэшированный объект, для записи лучше загрузить свежую копию
        # прямо перед обработкой. Это займет время, но гарантирует чистоту данных.
        with st.spinner("Подготовка файла для записи..."):
            wb_to_process = load_workbook(BytesIO(uploaded_file.getvalue()))
        
        progress_bar = st.progress(0)
        log_box = st.empty()
        
        total_sheets = len(selected_sheets)
        sheets_done = 0
        total_links_checked = 0
        slack_report = []

        session = requests.Session()
        session.headers.update(get_headers(api_key))

        # --- ОСНОВНОЙ ЦИКЛ ПО ЛИСТАМ ---
        for sheet_name in selected_sheets:
            log_box.markdown(f"⏳ **Лист: {sheet_name}** — подготовка данных...")
            
            ws = wb_to_process[sheet_name]
            header_row = find_header_row(ws)
            
            # Добавляем заголовок Index
            ws.cell(row=header_row, column=4).value = "Index"

            urls_map = {} # { url: [rows] }
            raw_urls = []
            
            # Сбор URL
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row=r, column=2).value
                if looks_like_url(val):
                    clean_url = val.strip()
                    raw_urls.append(clean_url)
                    if clean_url not in urls_map:
                        urls_map[clean_url] = []
                    urls_map[clean_url].append(r)
            
            if not raw_urls:
                log_box.warning(f"Лист {sheet_name}: ссылок не найдено.")
                sheets_done += 1
                progress_bar.progress(sheets_done / total_sheets)
                continue

            # Отправка задачи в API
            log_box.markdown(f"⏳ **Лист: {sheet_name}** — отправка {len(raw_urls)} ссылок в API...")
            
            try:
                # 1. Create Task
                create_resp = session.post(
                    f"{SPEEDY_BASE_URL}/task/google/checker/create",
                    json={"title": f"Streamlit {sheet_name}", "urls": raw_urls}
                )
                c_data = create_resp.json()
                
                if c_data.get("code") != 0:
                    log_box.error(f"Ошибка API на листе {sheet_name}: {c_data}")
                    slack_report.append(f"• List *{sheet_name}*: API Error")
                    continue
                
                task_id = c_data.get("task_id")
                
                # 2. Polling (ожидание)
                is_completed = False
                attempts = 0
                max_attempts = 100 # ~5 минут макс
                
                while not is_completed and attempts < max_attempts:
                    time.sleep(3)
                    st_resp = session.post(
                        f"{SPEEDY_BASE_URL}/task/google/checker/status",
                        json={"task_ids": [task_id]}
                    )
                    s_data = st_resp.json()
                    res_list = s_data.get("result", [])
                    
                    if not res_list:
                        break
                        
                    task_info = res_list[0]
                    
                    if task_info.get("is_completed"):
                        is_completed = True
                    else:
                        processed = task_info.get("processed_count", 0)
                        total_cnt = task_info.get("size", 0)
                        log_box.markdown(f"⏳ **Лист: {sheet_name}** — проверяем... ({processed}/{total_cnt})")
                        attempts += 1
                
                if not is_completed:
                    log_box.error(f"Таймаут проверки листа {sheet_name}")
                    slack_report.append(f"• List *{sheet_name}*: Timeout")
                    continue

                # 3. Get Report
                rep_resp = session.post(
                    f"{SPEEDY_BASE_URL}/task/google/checker/report",
                    json={"task_id": task_id}
                )
                r_data = rep_resp.json()
                indexed_links = set(r_data.get("result", {}).get("indexed_links", []))
                
                # 4. Запись в Excel
                log_box.markdown(f"💾 **Лист: {sheet_name}** — сохранение результатов...")
                
                for url, rows in urls_map.items():
                    # Простая проверка: есть ли URL в списке проиндексированных
                    is_indexed = url in indexed_links
                    
                    for r_idx in rows:
                        # Пишем TRUE / FALSE
                        ws.cell(row=r_idx, column=4).value = is_indexed

                count_idx = len(indexed_links)
                count_all = len(raw_urls)
                total_links_checked += count_all
                slack_report.append(f"• List *{sheet_name}*: {count_idx}/{count_all} indexed")

            except Exception as e:
                log_box.error(f"Exception on {sheet_name}: {e}")
                slack_report.append(f"• List *{sheet_name}*: Script Exception")
            
            sheets_done += 1
            progress_bar.progress(sheets_done / total_sheets)

        # --- ЗАВЕРШЕНИЕ ---
        log_box.success("✅ Все листы обработаны!")
        
        # Сохранение в буфер
        out_buffer = BytesIO()
        wb_to_process.save(out_buffer)
        out_buffer.seek(0)
        
        st.download_button(
            label="📥 Скачать результат (.xlsx)",
            data=out_buffer,
            file_name="speedy_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Slack
        if slack_report:
            header = f"🤖 *SpeedyIndex Check Report*\nTotal Links: {total_links_checked}\n\n"
            msg = header + "\n".join(slack_report)
            send_slack_notification(slack_token, slack_channel, msg)
            st.toast("Отчет отправлен в Slack!")
